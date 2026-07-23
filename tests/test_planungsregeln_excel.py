from pathlib import Path
import tempfile
import unittest

import openpyxl

from backend.db.sqlite_adapter import SQLiteAdapter
from backend.services.export_service import export_excel, export_planning_excel
from backend.services.import_service import _import_planungsregeln


def _database(path: Path) -> SQLiteAdapter:
    db = SQLiteAdapter(str(path))
    db.connect()
    db.create_tables()
    return db


class PlanungsregelnExcelTest(unittest.TestCase):
    def test_planning_proposal_is_import_compatible(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir)
            db = _database(path / "planning.db")
            try:
                db.execute(
                    "INSERT INTO intervalle (kuerzel, bedeutung) VALUES (?, ?)",
                    ("W", "Wöchentlich"),
                )
                db.execute(
                    "INSERT INTO termine (bespr_nr, bezeichnung, intervall, dauer_min) VALUES (?, ?, ?, ?)",
                    (44, "Planung", "W", 60),
                )
                proposal_path = path / "proposal.xlsx"
                export_planning_excel(
                    db,
                    str(proposal_path),
                    [{"woche": 3, "tag": "Tue", "start": "09:00", "bespr_nr": 44}],
                    3,
                    3,
                )
            finally:
                db.disconnect()

            workbook = openpyxl.load_workbook(proposal_path, read_only=True)
            try:
                row = list(workbook["Terminliste"].iter_rows(min_row=2, values_only=True))[0]
                self.assertEqual(row[:8], (3, "Tue", "09:00", "10:00", 44, "Planung", "W", 60))
                self.assertIn("Planungsregeln", workbook.sheetnames)
            finally:
                workbook.close()

    def test_planungsregeln_are_exported_as_dedicated_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            db = _database(temp_path / "source.db")
            try:
                db.execute(
                    "INSERT INTO planungsregeln "
                    "(bezeichnung, typ, bedingung, prioritaet, aktiv) VALUES (?, ?, ?, ?, ?)",
                    ("Arbeitszeit", "constraint", "Besprechungen nur von 08:30 bis 17:30 Uhr", 2, 1),
                )
                export_path = temp_path / "export.xlsx"
                export_excel(db, str(export_path))
            finally:
                db.disconnect()

            workbook = openpyxl.load_workbook(export_path, read_only=True)
            try:
                self.assertIn("Planungsregeln", workbook.sheetnames)
                rows = list(workbook["Planungsregeln"].iter_rows(values_only=True))
                self.assertEqual(rows[0], ("Nr.", "Bezeichnung", "Typ", "Regel", "Priorität", "Aktiv"))
                self.assertEqual(rows[1], (
                    1,
                    "Arbeitszeit",
                    "constraint",
                    "Besprechungen nur von 08:30 bis 17:30 Uhr",
                    2,
                    "Ja",
                ))
            finally:
                workbook.close()

    def test_planungsregeln_sheet_can_be_imported_and_numbered(self) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        self.assertIsNotNone(sheet)
        sheet.title = "Planungsregeln"
        sheet.append(["Nr.", "Bezeichnung", "Typ", "Regel", "Priorität", "Aktiv"])
        sheet.append([17, "Mittagspause", "exclusion", "Keine Meetings von 12:00 bis 13:00 Uhr", 1, "Ja"])
        sheet.append([18, "", "invalid", "Freitags nur bis 13:00 Uhr", 99, "Nein"])

        with tempfile.TemporaryDirectory() as temp_dir:
            db = _database(Path(temp_dir) / "target.db")
            try:
                _import_planungsregeln(db, sheet)
                rules = db.fetchall(
                    "SELECT id, bezeichnung, typ, bedingung, prioritaet, aktiv "
                    "FROM planungsregeln ORDER BY id"
                )
            finally:
                db.disconnect()

        self.assertEqual(rules, [
            {
                "id": 1,
                "bezeichnung": "Mittagspause",
                "typ": "exclusion",
                "bedingung": "Keine Meetings von 12:00 bis 13:00 Uhr",
                "prioritaet": 1,
                "aktiv": 1,
            },
            {
                "id": 2,
                "bezeichnung": "Regel 18",
                "typ": "constraint",
                "bedingung": "Freitags nur bis 13:00 Uhr",
                "prioritaet": 10,
                "aktiv": 0,
            },
        ])


if __name__ == "__main__":
    unittest.main()
