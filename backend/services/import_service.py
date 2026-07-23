"""Import service: reads an Excel file and populates the database."""

import logging
from typing import Any

import openpyxl

from backend.db.interface import DatabaseInterface

logger = logging.getLogger(__name__)

# Missing interval codes found in Termine but absent from the Intervalle sheet
EXTRA_INTERVALS = {
    "2M": "Alle 2 Monate",
    "W6": "Alle 6 Wochen",
}


def import_excel(db: DatabaseInterface, filepath: str, allow_invalid: bool = False) -> None:
    """Import all sheets from the Excel file into the database.

    Clears existing data first, then imports in FK-safe order.
    """
    logger.info("Starting import from %s", filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True)

    try:
        issues = _validate_workbook(wb)
        if issues and not allow_invalid:
            raise ValueError("Import enthält ungültige Referenzen: " + "; ".join(issue["message"] for issue in issues[:10]))
        db.clear_tables()
        _import_bereiche(db, wb["Bereiche"])
        _import_usergruppen(db, wb["Usergruppen"])
        _import_intervalle(db, wb["Intervalle"])
        _import_termine(db, wb["Termine"], skip_invalid=allow_invalid)
        _import_terminliste(db, wb["Terminliste"], skip_invalid=allow_invalid)
        if "Planungsregeln" in wb.sheetnames:
            _import_planungsregeln(db, wb["Planungsregeln"])
        _import_optional_resources(db, wb)
    finally:
        wb.close()

    logger.info("Import completed successfully")


def validate_excel(filepath: str) -> list[dict]:
    """Return all missing or illogical workbook references before destructive import."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    try:
        return _validate_workbook(wb)
    finally:
        wb.close()


def _rows(wb: Any, sheet: str) -> list[tuple]:
    return list(wb[sheet].iter_rows(min_row=2, values_only=True)) if sheet in wb.sheetnames else []


def _validate_workbook(wb: Any) -> list[dict]:
    issues: list[dict] = []
    required = {"Bereiche", "Usergruppen", "Intervalle", "Termine", "Terminliste"}
    for sheet in sorted(required - set(wb.sheetnames)):
        issues.append({"severity": "error", "sheet": sheet, "message": f"Pflicht-Tabellenblatt fehlt: {sheet}"})
    if issues:
        return issues
    areas = {str(row[0]).strip() for row in _rows(wb, "Bereiche") if row and row[0] is not None}
    groups = {str(row[1]).strip() for row in _rows(wb, "Usergruppen") if len(row) > 1 and row[1] is not None}
    intervals = {str(row[0]).strip() for row in _rows(wb, "Intervalle") if row and row[0] is not None} | set(EXTRA_INTERVALS)
    meetings = {int(row[0]) for row in _rows(wb, "Termine") if row and row[0] is not None}
    for index, row in enumerate(_rows(wb, "Usergruppen"), 2):
        if row and row[0] is not None and str(row[0]).strip() not in areas:
            issues.append({"severity": "error", "sheet": "Usergruppen", "row": index, "message": f"Bereich {row[0]} existiert nicht"})
    for index, row in enumerate(_rows(wb, "Termine"), 2):
        if not row or row[0] is None:
            continue
        if len(row) < 3 or str(row[2]).strip() not in intervals:
            issues.append({"severity": "error", "sheet": "Termine", "row": index, "message": f"Intervall {row[2] if len(row) > 2 else ''} existiert nicht"})
        for value in row[4:]:
            if value is not None and str(value).strip() and str(value).strip() not in groups:
                issues.append({"severity": "error", "sheet": "Termine", "row": index, "message": f"Teilnehmergruppe {value} existiert nicht"})
    for index, row in enumerate(_rows(wb, "Terminliste"), 2):
        if len(row) > 4 and row[4] is not None and int(row[4]) not in meetings:
            issues.append({"severity": "error", "sheet": "Terminliste", "row": index, "message": f"Meeting {row[4]} existiert nicht"})
    people = {int(row[0]) for row in _rows(wb, "Personen") if row and row[0] is not None}
    roles = {int(row[0]) for row in _rows(wb, "Rollen") if row and row[0] is not None}
    rooms = {int(row[0]) for row in _rows(wb, "Raeume") if row and row[0] is not None}
    components = {int(row[0]) for row in _rows(wb, "Komponenten") if row and row[0] is not None}
    reference_sheets = [
        ("RollenPersonen", roles, people, "Rolle", "Person"),
        ("RollenGruppen", roles, groups, "Rolle", "Gruppe"),
        ("TerminRaeume", meetings, rooms, "Termin", "Raum"),
        ("TerminKomponenten", meetings, components, "Termin", "Komponente"),
    ]
    for sheet, left_set, right_set, left_name, right_name in reference_sheets:
        for index, row in enumerate(_rows(wb, sheet), 2):
            if len(row) < 2 or row[0] is None or row[1] is None:
                continue
            left = int(row[0])
            right = str(row[1]).strip() if sheet == "RollenGruppen" else int(row[1])
            if left not in left_set:
                issues.append({"severity": "error", "sheet": sheet, "row": index, "message": f"{left_name} {left} existiert nicht"})
            if right not in right_set:
                issues.append({"severity": "error", "sheet": sheet, "row": index, "message": f"{right_name} {right} existiert nicht"})
    return issues


def _import_bereiche(db: DatabaseInterface, ws: Any) -> None:
    """Import the Bereiche (departments) sheet."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data = [(r[0], r[1]) for r in rows if r[0] is not None]
    db.executemany("INSERT INTO bereiche (gruppe, bereich) VALUES (?, ?)", data)
    logger.info("Imported %d bereiche", len(data))


def _import_usergruppen(db: DatabaseInterface, ws: Any) -> None:
    """Import the Usergruppen (user groups) sheet."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data = []
    for r in rows:
        if r[0] is None or r[1] is None:
            continue
        bereich = str(r[0]).strip()
        nummer = str(r[1]).strip()
        bezeichnung = str(r[2]).strip() if r[2] else ""
        name = str(r[3]).strip() if r[3] else None
        data.append((nummer, bereich, bezeichnung, name))
    db.executemany(
        "INSERT INTO usergruppen (nummer, bereich, bezeichnung, name) VALUES (?, ?, ?, ?)",
        data,
    )
    logger.info("Imported %d usergruppen", len(data))


def _import_intervalle(db: DatabaseInterface, ws: Any) -> None:
    """Import the Intervalle sheet plus any missing interval codes."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data = []
    seen = set()
    for r in rows:
        if r[0] is None:
            continue
        kuerzel = str(r[0]).strip()
        bedeutung = str(r[1]).strip() if r[1] else ""
        data.append((kuerzel, bedeutung))
        seen.add(kuerzel)

    # Add missing intervals discovered during planning
    for k, v in EXTRA_INTERVALS.items():
        if k not in seen:
            data.append((k, v))
            logger.debug("Added missing interval: %s = %s", k, v)

    db.executemany("INSERT INTO intervalle (kuerzel, bedeutung) VALUES (?, ?)", data)
    logger.info("Imported %d intervalle", len(data))


def _import_termine(db: DatabaseInterface, ws: Any, skip_invalid: bool = False) -> None:
    """Import the Termine (meetings) sheet and participant junction table."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    termine_data = []
    teilnehmer_data = []

    for r in rows:
        if r[0] is None:
            continue
        bespr_nr = int(r[0])
        bezeichnung = str(r[1]).strip() if r[1] else ""
        intervall = str(r[2]).strip() if r[2] else ""
        if skip_invalid and not db.fetchone("SELECT kuerzel FROM intervalle WHERE kuerzel = ?", (intervall,)):
            logger.warning("Skipping meeting %s with unknown interval %s", bespr_nr, intervall)
            continue
        dauer_min = int(r[3]) if r[3] else 0
        termine_data.append((bespr_nr, bezeichnung, intervall, dauer_min))

        # Columns E onwards (index 4+) are participant codes
        seen_codes = set()
        for col_idx in range(4, len(r)):
            val = r[col_idx]
            if val is not None:
                code = str(val).strip()
                if code and code not in seen_codes and (
                    not skip_invalid or db.fetchone("SELECT nummer FROM usergruppen WHERE nummer = ?", (code,))
                ):
                    seen_codes.add(code)
                    teilnehmer_data.append((bespr_nr, code))

    db.executemany(
        "INSERT INTO termine (bespr_nr, bezeichnung, intervall, dauer_min) VALUES (?, ?, ?, ?)",
        termine_data,
    )
    logger.info("Imported %d termine", len(termine_data))

    db.executemany(
        "INSERT INTO termin_teilnehmer (bespr_nr, usergruppe) VALUES (?, ?)",
        teilnehmer_data,
    )
    logger.info("Imported %d termin_teilnehmer entries", len(teilnehmer_data))


def _import_terminliste(db: DatabaseInterface, ws: Any, skip_invalid: bool = False) -> None:
    """Import the Terminliste sheet (only non-derived columns)."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data = []
    for r in rows:
        # Filter out None/empty rows (e.g. trailing row 828)
        if r[0] is None or r[4] is None:
            continue
        woche = int(r[0])
        tag = str(r[1]).strip()
        start = str(r[2]).strip()
        meeting = int(r[4])
        if skip_invalid and not db.fetchone("SELECT bespr_nr FROM termine WHERE bespr_nr = ?", (meeting,)):
            logger.warning("Skipping schedule row for unknown meeting %s", meeting)
            continue
        data.append((woche, tag, start, meeting))

    db.executemany(
        "INSERT INTO terminliste (woche, tag, start, meeting) VALUES (?, ?, ?, ?)",
        data,
    )
    logger.info("Imported %d terminliste entries", len(data))


def _import_planungsregeln(db: DatabaseInterface, ws: Any) -> None:
    """Import planning rules; their IDs are assigned automatically in sheet order."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data = []
    for r in rows:
        if len(r) < 4 or r[3] is None:
            continue
        exported_number = r[0] if len(r) > 0 else None
        bezeichnung = str(r[1]).strip() if len(r) > 1 and r[1] else (
            f"Regel {exported_number}" if exported_number is not None else "Regel"
        )
        typ = str(r[2]).strip() if len(r) > 2 and r[2] else "constraint"
        if typ not in {"constraint", "preference", "exclusion", "requirement"}:
            typ = "constraint"
        bedingung = str(r[3]).strip()
        try:
            prioritaet = max(1, min(10, int(r[4]))) if len(r) > 4 and r[4] is not None else 5
        except (TypeError, ValueError):
            prioritaet = 5
        aktiv_value = r[5] if len(r) > 5 else True
        aktiv = 0 if str(aktiv_value).strip().lower() in {"0", "nein", "no", "false"} else 1
        data.append((bezeichnung, typ, bedingung, prioritaet, aktiv))

    db.executemany(
        "INSERT INTO planungsregeln "
        "(bezeichnung, typ, bedingung, prioritaet, aktiv) VALUES (?, ?, ?, ?, ?)",
        data,
    )
    logger.info("Imported %d planungsregeln", len(data))


def _import_optional_resources(db: DatabaseInterface, wb: Any) -> None:
    """Import extended resource sheets and only keep resolvable assignments."""
    if "Personen" in wb.sheetnames:
        data = []
        for row in _rows(wb, "Personen"):
            if row and row[0] is not None:
                data.append((int(row[0]), str(row[1] or ""), str(row[2] or ""), str(row[3] or ""),
                             str(row[4] or ""), str(row[5] or ""), row[6] or None, int(row[7] or 0)))
        db.executemany(
            "INSERT INTO personen (id, vorname, nachname, email, telefon, standort, usergruppe, aktiv) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            data,
        )
    if "Rollen" in wb.sheetnames:
        db.executemany(
            "INSERT INTO rollen (id, bezeichnung, beschreibung, farbe) VALUES (?, ?, ?, ?)",
            [(int(r[0]), str(r[1] or ""), str(r[2] or ""), str(r[3] or "#71717a"))
             for r in _rows(wb, "Rollen") if r and r[0] is not None],
        )
    if "Raeume" in wb.sheetnames:
        db.executemany(
            "INSERT INTO raeume (id, bezeichnung, gebaeude, kapazitaet, ausstattung, aktiv) VALUES (?, ?, ?, ?, ?, ?)",
            [(int(r[0]), str(r[1] or ""), str(r[2] or ""), int(r[3] or 0), str(r[4] or ""), int(r[5] or 0))
             for r in _rows(wb, "Raeume") if r and r[0] is not None],
        )
    if "Komponenten" in wb.sheetnames:
        db.executemany(
            "INSERT INTO komponenten (id, bezeichnung, typ, beschreibung, verfuegbar) VALUES (?, ?, ?, ?, ?)",
            [(int(r[0]), str(r[1] or ""), str(r[2] or ""), str(r[3] or ""), int(r[4] or 0))
             for r in _rows(wb, "Komponenten") if r and r[0] is not None],
        )
    mappings = [
        ("RollenPersonen", "person_rollen", "rolle_id", "person_id", "rollen", "id", "personen", "id"),
        ("RollenGruppen", "rolle_gruppen", "rolle_id", "usergruppe", "rollen", "id", "usergruppen", "nummer"),
        ("TerminRaeume", "termin_raeume", "bespr_nr", "raum_id", "termine", "bespr_nr", "raeume", "id"),
        ("TerminKomponenten", "termin_komponenten", "bespr_nr", "komponente_id", "termine", "bespr_nr", "komponenten", "id"),
    ]
    for sheet, table, left_col, right_col, left_table, left_key, right_table, right_key in mappings:
        if sheet not in wb.sheetnames:
            continue
        valid = []
        for row in _rows(wb, sheet):
            if len(row) < 2 or row[0] is None or row[1] is None:
                continue
            left = int(row[0])
            right = str(row[1]).strip() if sheet == "RollenGruppen" else int(row[1])
            if db.fetchone(f"SELECT {left_key} FROM {left_table} WHERE {left_key} = ?", (left,)) and \
               db.fetchone(f"SELECT {right_key} FROM {right_table} WHERE {right_key} = ?", (right,)):
                valid.append((left, right))
        db.executemany(f"INSERT INTO {table} ({left_col}, {right_col}) VALUES (?, ?)", valid)
