"""Export service: writes the database contents to an Excel file."""

import logging

import openpyxl
from openpyxl.utils import get_column_letter

from backend.db.interface import DatabaseInterface

logger = logging.getLogger(__name__)


def export_excel(db: DatabaseInterface, filepath: str) -> None:
    """Export all database tables to an Excel file matching the original format."""
    logger.info("Starting export to %s", filepath)
    wb = openpyxl.Workbook()

    # Remove default sheet
    if wb.active is not None:
        wb.remove(wb.active)

    _export_bereiche(db, wb)
    _export_usergruppen(db, wb)
    _export_termine(db, wb)
    _export_terminliste(db, wb)
    _export_intervalle(db, wb)
    _export_planungsregeln(db, wb)
    _export_personen(db, wb)
    _export_rollen(db, wb)
    _export_raeume(db, wb)
    _export_komponenten(db, wb)
    _export_rollenzuordnungen(db, wb)
    _export_terminressourcen(db, wb)

    wb.save(filepath)
    logger.info("Export completed: %s", filepath)


def export_planning_excel(
    db: DatabaseInterface,
    filepath: str,
    proposals: list[dict],
    week_from: int,
    week_to: int,
) -> None:
    """Export a normal workbook whose Terminliste contains the AI proposal."""
    export_excel(db, filepath)
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook["Terminliste"]
    retained = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)
                if row[0] is not None and not week_from <= int(row[0]) <= week_to]
    sheet.delete_rows(2, max(0, sheet.max_row - 1))
    for row in retained:
        sheet.append(row)
    for proposal in proposals:
        try:
            meeting_number = int(proposal["bespr_nr"])
            week = int(proposal["woche"])
            day = str(proposal["tag"])
            start = str(proposal["start"])
        except (KeyError, TypeError, ValueError):
            continue
        if week < week_from or week > week_to or day not in {"Mon", "Tue", "Wed", "Thu", "Fri"}:
            continue
        meeting = db.fetchone(
            "SELECT bezeichnung, intervall, dauer_min FROM termine WHERE bespr_nr = ?",
            (meeting_number,),
        )
        if not meeting:
            continue
        participants = db.fetchall(
            "SELECT usergruppe FROM termin_teilnehmer WHERE bespr_nr = ? ORDER BY usergruppe",
            (meeting_number,),
        )
        sheet.append([
            week,
            day,
            start,
            _add_minutes(start, meeting["dauer_min"]),
            meeting_number,
            meeting["bezeichnung"],
            meeting["intervall"],
            meeting["dauer_min"],
            ", ".join(row["usergruppe"] for row in participants),
        ])
    workbook.save(filepath)


def _export_bereiche(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    """Export Bereiche sheet."""
    ws = wb.create_sheet("Bereiche")
    ws.append(["Gruppe", "Bereich"])
    rows = db.fetchall("SELECT gruppe, bereich FROM bereiche ORDER BY gruppe")
    for r in rows:
        ws.append([r["gruppe"], r["bereich"]])
    logger.info("Exported %d bereiche", len(rows))


def _export_usergruppen(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    """Export Usergruppen sheet."""
    ws = wb.create_sheet("Usergruppen")
    ws.append(["Bereich", "Nummer", "Bezeichnung", "Name"])
    rows = db.fetchall(
        "SELECT bereich, nummer, bezeichnung, name FROM usergruppen ORDER BY nummer"
    )
    for r in rows:
        ws.append([r["bereich"], r["nummer"], r["bezeichnung"], r["name"]])
    logger.info("Exported %d usergruppen", len(rows))


def _export_termine(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    """Export Termine sheet with variable-width participant columns."""
    ws = wb.create_sheet("Termine")

    # Determine max number of participants across all meetings
    max_result = db.fetchone(
        "SELECT MAX(cnt) as max_cnt FROM "
        "(SELECT COUNT(*) as cnt FROM termin_teilnehmer GROUP BY bespr_nr)"
    )
    max_participants = (
        max_result["max_cnt"] if max_result and max_result["max_cnt"] else 0
    )

    # Header: first 4 fixed columns + participant columns
    header = ["Bespr.Nr.", "Bezeichung", "Intervall", "Dauer (min)"]
    if max_participants > 0:
        header.append("Usergruppen zur Teilnahme")
        # Remaining participant columns have no header
        header.extend([""] * (max_participants - 1))
    ws.append(header)

    # Data rows
    termine = db.fetchall(
        "SELECT bespr_nr, bezeichnung, intervall, dauer_min FROM termine ORDER BY bespr_nr"
    )
    for t in termine:
        participants = db.fetchall(
            "SELECT usergruppe FROM termin_teilnehmer WHERE bespr_nr = ? ORDER BY usergruppe",
            (t["bespr_nr"],),
        )
        row = [t["bespr_nr"], t["bezeichnung"], t["intervall"], t["dauer_min"]]
        row.extend(p["usergruppe"] for p in participants)
        ws.append(row)

    logger.info("Exported %d termine", len(termine))


def _add_minutes(time_str: str, minutes: int) -> str:
    """Add minutes to a HH:MM time string and return HH:MM."""
    parts = time_str.split(":")
    total_min = int(parts[0]) * 60 + int(parts[1]) + minutes
    h = total_min // 60
    m = total_min % 60
    return f"{h:02d}:{m:02d}"


def _export_terminliste(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    """Export Terminliste sheet with all derived fields reconstructed."""
    ws = wb.create_sheet("Terminliste")
    ws.append(
        [
            "Woche",
            "Tag",
            "Start",
            "Ende",
            "Meeting",
            "Name",
            "Intervall",
            "Dauer (min)",
            "Teilnehmer (Quelle)",
        ]
    )

    rows = db.fetchall("""
        SELECT
            tl.woche,
            tl.tag,
            tl.start,
            t.bespr_nr,
            t.bezeichnung,
            t.intervall,
            t.dauer_min
        FROM terminliste tl
        JOIN termine t ON tl.meeting = t.bespr_nr
        ORDER BY tl.woche, 
            CASE tl.tag 
                WHEN 'Mon' THEN 1 
                WHEN 'Tue' THEN 2 
                WHEN 'Wed' THEN 3 
                WHEN 'Thu' THEN 4 
                WHEN 'Fri' THEN 5 
            END,
            tl.start,
            tl.id
    """)

    for r in rows:
        # Get participants for this meeting
        participants = db.fetchall(
            "SELECT usergruppe FROM termin_teilnehmer WHERE bespr_nr = ? ORDER BY usergruppe",
            (r["bespr_nr"],),
        )
        teilnehmer_str = ", ".join(p["usergruppe"] for p in participants)
        ende = _add_minutes(r["start"], r["dauer_min"])

        ws.append(
            [
                r["woche"],
                r["tag"],
                r["start"],
                ende,
                r["bespr_nr"],
                r["bezeichnung"],
                r["intervall"],
                r["dauer_min"],
                teilnehmer_str,
            ]
        )

    logger.info("Exported %d terminliste entries", len(rows))


def _export_intervalle(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    """Export Intervalle sheet."""
    ws = wb.create_sheet("Intervalle")
    ws.append(["Kürzel", "Bedeutung"])
    rows = db.fetchall("SELECT kuerzel, bedeutung FROM intervalle ORDER BY kuerzel")
    for r in rows:
        ws.append([r["kuerzel"], r["bedeutung"]])
    logger.info("Exported %d intervalle", len(rows))


def _export_planungsregeln(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    """Export editable planning rules as a dedicated worksheet."""
    ws = wb.create_sheet("Planungsregeln")
    ws.append(["Nr.", "Bezeichnung", "Typ", "Regel", "Priorität", "Aktiv"])
    rows = db.fetchall(
        "SELECT id, bezeichnung, typ, bedingung, prioritaet, aktiv "
        "FROM planungsregeln ORDER BY id"
    )
    for r in rows:
        ws.append([
            r["id"],
            r["bezeichnung"],
            r["typ"],
            r["bedingung"],
            r["prioritaet"],
            "Ja" if r["aktiv"] else "Nein",
        ])
    ws.freeze_panes = "A2"
    widths = (10, 28, 18, 100, 12, 10)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    logger.info("Exported %d planungsregeln", len(rows))


def _export_personen(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Personen")
    ws.append(["ID", "Vorname", "Nachname", "E-Mail", "Telefon", "Standort", "Usergruppe", "Aktiv"])
    for row in db.fetchall("SELECT id, vorname, nachname, email, telefon, standort, usergruppe, aktiv FROM personen ORDER BY id"):
        ws.append([row["id"], row["vorname"], row["nachname"], row["email"], row["telefon"], row["standort"], row["usergruppe"], row["aktiv"]])


def _export_rollen(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Rollen")
    ws.append(["ID", "Bezeichnung", "Beschreibung", "Farbe"])
    for row in db.fetchall("SELECT id, bezeichnung, beschreibung, farbe FROM rollen ORDER BY id"):
        ws.append([row["id"], row["bezeichnung"], row["beschreibung"], row["farbe"]])


def _export_raeume(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Raeume")
    ws.append(["ID", "Bezeichnung", "Gebäude", "Kapazität", "Ausstattung", "Aktiv"])
    for row in db.fetchall("SELECT id, bezeichnung, gebaeude, kapazitaet, ausstattung, aktiv FROM raeume ORDER BY id"):
        ws.append([row["id"], row["bezeichnung"], row["gebaeude"], row["kapazitaet"], row["ausstattung"], row["aktiv"]])


def _export_komponenten(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Komponenten")
    ws.append(["ID", "Bezeichnung", "Typ", "Beschreibung", "Verfügbar"])
    for row in db.fetchall("SELECT id, bezeichnung, typ, beschreibung, verfuegbar FROM komponenten ORDER BY id"):
        ws.append([row["id"], row["bezeichnung"], row["typ"], row["beschreibung"], row["verfuegbar"]])


def _export_rollenzuordnungen(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("RollenPersonen")
    ws.append(["Rolle-ID", "Person-ID"])
    for row in db.fetchall("SELECT rolle_id, person_id FROM person_rollen ORDER BY rolle_id, person_id"):
        ws.append([row["rolle_id"], row["person_id"]])
    ws = wb.create_sheet("RollenGruppen")
    ws.append(["Rolle-ID", "Usergruppe"])
    for row in db.fetchall("SELECT rolle_id, usergruppe FROM rolle_gruppen ORDER BY rolle_id, usergruppe"):
        ws.append([row["rolle_id"], row["usergruppe"]])


def _export_terminressourcen(db: DatabaseInterface, wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("TerminRaeume")
    ws.append(["Bespr.Nr.", "Raum-ID"])
    for row in db.fetchall("SELECT bespr_nr, raum_id FROM termin_raeume ORDER BY bespr_nr, raum_id"):
        ws.append([row["bespr_nr"], row["raum_id"]])
    ws = wb.create_sheet("TerminKomponenten")
    ws.append(["Bespr.Nr.", "Komponente-ID"])
    for row in db.fetchall("SELECT bespr_nr, komponente_id FROM termin_komponenten ORDER BY bespr_nr, komponente_id"):
        ws.append([row["bespr_nr"], row["komponente_id"]])
